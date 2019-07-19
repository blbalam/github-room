from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from smtplib import SMTP_SSL

SMTP_SERVER = 'smtp.naver.com'
SMTP_PORT = 465
SMTP_USER = 'lthlovelee'
# 실제 비밀번호를 입력해야 합니다.
SMTP_PASSWORD = ''


#alternative(텍스트 메일을 발송할 때 필요한 키워드)
#mixed(첨부파일이 포함된 메일을 발송할 때 필요한 키워드)
#digest(메일을 전달할 때 필요한  키워드)

def send_mail(name, addr, contents, attachment=False):
    msg = MIMEMultipart('alternative')

    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = 'lthlovelee <%s>'%SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, 메일이 도착했습니다.'

    text = MIMEText(contents)
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octet-stream')
        f = open(attachment, 'rb')
        file_contents = f.read()
        file_data.set_payload(file_contents)
        encoders.encode_base64(file_data)

        from os.path import basename
        filename = basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(file_data)

    smtp = SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()


#엑셀로부터 데이터 읽어오기
from my_email import send_mail
from openpyxl import load_workbook

xlsx = load_workbook('수강생_결제정보.xlsx', read_only=True)
sheet = xlsx.active

for row in sheet.iter_rows():
    name = row[0].value
    mail = row[1].value
    status = row[3].value

    if status == '결제완료':
        contents = '결제완료가 확인되어 커리큘럼을 안내해드립니다.'
        send_mail(name, mail, contents, '커리큘럼.xlsx')