from os import listdir
from openpyxl import load_workbook, Workbook

'''단일셀 데이터 읽기
result.xlsx를 읽은 후 기본 시트를 선택
xlsx = load_workbook('result.xlsx', read_only=True)
sheet = xlsx.active

A1 셀의 데이터를 출력
print(sheet['A1'].value)
B1 셀의 데이터를 출력
print(sheet['B1'].value)

A2,A3,B2,B3 셀을 가져옴
rows = sheet['A2:B3']
각 행을 가져오기 위한 반복문
for row in rows:
    # 각 행의 셀을 가져오기 위한 반복문
    for cell in row:
        print(cell.value)
'''
'''단일셀 데이터 쓰기
새로운 엑셀 파일 생성을 위한 클래스 변수 생성
xlsx = Workbook()
sheet = xlsx.active

A1 셀에 문자열 값 추가
sheet['A1'] = 'my input data'
엑셀 파일로 저장
xlsx.save('other.xlsx')
'''
'''행단위 셀 데이터 쓰기
첫번째 행에 데이터 추가
sheet.append(['A1-data', 'B1-data', 'C1-data'])
두번째 행에 데이터 추가
sheet.append(['A2-data', 'B2-data', 'C2-data'])
엑셀 파일로 저장
xlsx.save('other.xlsx')
'''

'''특정 행과 열의 셀데이터 읽기
result.xlsx를 읽은 후 기본 시트를 선택
xlsx = load_workbook('result.xlsx', read_only=True)
sheet = xlsx.active

첫번째 행을 가져옴
row = sheet['1']
첫번째 행의 각 컬럼의 값을 출력
for cell in row:
    print(cell.value)
    
# 첫번째 열을 가져옴
col = sheet['A']
# 첫번째 열의 각 행 값을 출력
for cell in col:
    print(cell.value)
'''

'''여러 행과 옇 의 셀 데이터 읽기
result.xlsx를 읽은 후 기본 시트를 선택
xlsx = load_workbook('result.xlsx', read_only=True)
sheet = xlsx.active

1~2 행을 가져옴
rows = sheet['1:2']
각 행을 가져오기 위한 반복문
for row in rows:
    # 각 행의 셀을 가져오기 위한 반복문
    for cell in row:
        print(cell.value)
        
A~B 열을 가져옴
cols = sheet['A:B']
각 열을 가져오기 위한 반복문
for col in cols:
    # 각 열의 셀을 가져오기 위한 반복문
    for cell in col:
        print(cell.value)
'''
'''새로운 시트 생성 저장하기
from openpyxl import Workbook

새로운 엑셀 파일생성을 위한 클래스 변수 생성
xlsx = Workbook()

새로운시트2 이름을 가진 시트생성
sheet = xlsx.create_sheet('새로운시트2')
sheet['A1'] = '데이터'

엑셀파일 저장
xlsx.save('other.xlsx')
'''
'''시트 지정하여 저장하기
from openpyxl import Workbook

새로운 엑셀 파일생성을 위한 클래스 변수 생성
xlsx = load_Workbook('other.xlsx')

새로운시트2 이름을 가진 시트생성
sheet = xlsx['새로운시트2']
print(sheet['A1'].value)

엑셀파일 저장
xlsx.save('other.xlsx')
'''

#폴더의 파일 목록 조회하기
files = listdir('.')


#새 엑셀 파일의 클래스 변수 생성하기
result_xlsx = Workbook()
result_sheet = result_xlsx.active

#xlsx 확장자가 아닌 파일 건너뛰기
for myfile in files:
    if myfile[-4:] != 'xlsx':
        continue

#엑셀파일 열기 및 시트 선택하기
    tg_xlsx = load_workbook(myfile, read_only=True)
    tg_sheet = tg_xlsx.active
    #활성화된 시트가 아니라 다른 시트를 가져오고 싶다면 ( tg_sheet = tg_xlsx["시트명"]


#데이터를 가져와 새 엑셀파일에 쓰기
    for row in tg_sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)

        result_sheet.append(row_data)
#새 엑셀파일 저장하기
result_xlsx.save('result.xlsx')
